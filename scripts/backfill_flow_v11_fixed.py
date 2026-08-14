import io,json
from openpyxl import load_workbook
import backfill_flow_v11 as m

def hdr_token(x):
    s=str(x or '').strip();n=m.nkey(s)
    if 'khoiluong' in n or n=='kl':return 'KL'
    if 'giatri' in n or n=='gt' or 'value' in n:return 'GT'
    return s

def export_rows(sym,kind):
    q=m.urlencode({'Type':'EXPORT','Symbol':sym,'Exchange':'HOSE','StartDate':'01/01/2018','EndDate':'31/12/2026','PageIndex':1,'PageSize':20})
    raw,ct=m.get_bytes(m.BASE+m.endpoint(kind)+'?'+q)
    if not raw.startswith(b'PK'):return []
    wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True);best=[]
    for ws in wb.worksheets:
        vals=list(ws.iter_rows(values_only=True))
        for hi in range(min(25,max(0,len(vals)-1))):
            p=[str(x or '').strip() for x in vals[hi]];c=[str(x or '').strip() for x in vals[hi+1]]
            joined=' '.join(m.nkey(x) for x in p+c)
            if not (('ngay' in joined or 'date' in joined) and any(x in joined for x in ['mua','buy']) and any(x in joined for x in ['ban','sell'])):continue
            width=max(len(p),len(c));headers=[];parent=''
            for j in range(width):
                a=p[j] if j<len(p) else '';b=c[j] if j<len(c) else ''
                if a:parent=a
                pn=m.nkey(parent);bn=m.nkey(b)
                if b and pn not in {'ma','mack','symbol','ngay','date'}:
                    headers.append((parent+' '+hdr_token(b)).strip())
                elif a:
                    headers.append(a)
                elif b:
                    headers.append(b)
                else:
                    headers.append(f'c{j}')
            # Only treat the next row as a second header when it actually contains
            # KL/GT-style subcolumns. Otherwise data starts immediately after hi.
            has_child=any(x in ' '.join(m.nkey(y) for y in c) for x in ['khoiluong','giatri','volume','value'])
            start=hi+2 if has_child else hi+1
            rows=[]
            for rr in vals[start:]:
                if not any(x is not None and str(x).strip() for x in rr):continue
                rows.append({headers[j] if j<len(headers) else f'c{j}':rr[j] for j in range(len(rr))})
            if len(rows)>len(best):best=rows
    return best

m.export_rows=export_rows
m.main()

import json,re,math,os,io
from pathlib import Path
from datetime import datetime,timezone
from concurrent.futures import ThreadPoolExecutor,as_completed
from urllib.parse import urlencode
from urllib.request import Request,urlopen
from openpyxl import load_workbook
import numpy as np

ROOT=Path(os.environ.get('GITHUB_WORKSPACE','.'));VERSION='VMEWS-FLOW-11.2.0';START='01/01/2018';END='31/12/2026';BASE='https://cafef.vn/du-lieu/Ajax/PageNew/DataHistory/'
UA={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/144 Safari/537.36','Accept':'*/*','Referer':'https://cafef.vn/du-lieu/lich-su-giao-dich.chn'}

def nkey(x):
 s=str(x or '').lower();s=s.translate(str.maketrans('áàạảãâấầậẩẫăắằặẳẵéèẹẻẽêếềệểễíìịỉĩóòọỏõôốồộổỗơớờợởỡúùụủũưứừựửữýỳỵỷỹđ','aaaaaaaaaaaaaaaaaeeeeeeeeeeeiiiiiooooooooooooooooouuuuuuuuuuuyyyyyd'));return re.sub(r'[^a-z0-9]','',s)
def asnum(x):
 if x is None:return 0.
 if isinstance(x,(int,float)):return float(x) if math.isfinite(float(x)) else 0.
 s=str(x).strip().replace('\xa0','').replace(' ','').replace('%','')
 if not s or s in {'-','--','nan','None'}:return 0.
 if ',' in s and '.' not in s:
  tail=s.rsplit(',',1)[-1];s=s.replace(',','.') if len(tail)<=3 else s.replace(',','')
 elif ',' in s and '.' in s:
  s=s.replace('.','').replace(',','.') if s.rfind(',')>s.rfind('.') else s.replace(',','')
 try:return float(s)
 except:return 0.
def get(url,timeout=20):
 with urlopen(Request(url,headers=UA),timeout=timeout) as r:return r.read(),str(r.headers.get('Content-Type') or '')
def date_iso(x):
 if hasattr(x,'date'):
  try:return x.date().isoformat()
  except:pass
 s=str(x or '').strip()
 for f in ('%d/%m/%Y','%d-%m-%Y','%Y-%m-%d','%d/%m/%y'):
  try:return datetime.strptime(s[:10],f).date().isoformat()
  except:pass
 return s[:10] if re.match(r'^\d{4}-\d{2}-\d{2}',s) else None
def combined_headers(ws,hi):
 top=[ws.cell(hi,c).value for c in range(1,ws.max_column+1)];bot=[ws.cell(hi+1,c).value for c in range(1,ws.max_column+1)] if hi<ws.max_row else [None]*len(top);sub=sum(any(w in nkey(x) for w in ('khoiluong','giatri','volume','value')) for x in bot if x is not None)>=2;out=[];parent=''
 for a,b in zip(top,bot):
  if a not in (None,''):parent=str(a)
  child=str(b) if sub and b not in (None,'') else ''
  if child and parent and nkey(child)!=nkey(parent):out.append(parent+' '+child)
  elif a not in (None,''):out.append(str(a))
  elif child:out.append(child)
  else:out.append(parent)
 return out,hi+(2 if sub else 1)
def export_rows(sym,kind):
 ep='GDKhoiNgoai.ashx' if kind=='foreign' else 'GDTuDoanh.ashx';q=urlencode({'Type':'EXPORT','Symbol':sym,'Exchange':'HOSE','StartDate':START,'EndDate':END,'PageIndex':1,'PageSize':20});raw,ct=get(BASE+ep+'?'+q,30)
 if not raw.startswith(b'PK'):return []
 try:ws=load_workbook(io.BytesIO(raw),read_only=True,data_only=True).active
 except:return []
 hi=None
 for r in range(1,min(ws.max_row,30)+1):
  if any('ngay' in nkey(ws.cell(r,c).value) for c in range(1,ws.max_column+1)):hi=r;break
 if hi is None:return []
 hdr,start=combined_headers(ws,hi);keys=[nkey(x) for x in hdr]
 def v(row,*groups):
  for group in groups:
   for i,k in enumerate(keys):
    if all(term in k for term in group):return asnum(row[i] if i<len(row) else 0)
  return 0.
 out=[]
 for vals in ws.iter_rows(min_row=start,values_only=True):
  d=None
  for i,k in enumerate(keys):
   if 'ngay' in k:d=date_iso(vals[i] if i<len(vals) else None);break
  if not d:continue
  if kind=='foreign':z={'date':d,'foreignBuyValue':v(vals,('mua','giatri'),('mua','gt'),('buy','value')),'foreignSellValue':v(vals,('ban','giatri'),('ban','gt'),('sell','value')),'foreignNetValue':v(vals,('rong','giatri'),('rong','gt'),('net','value')),'foreignBuyVolume':v(vals,('mua','khoiluong'),('mua','kl'),('buy','volume')),'foreignSellVolume':v(vals,('ban','khoiluong'),('ban','kl'),('sell','volume')),'foreignNetVolume':v(vals,('rong','khoiluong'),('rong','kl'),('net','volume')),'foreignRoom':v(vals,('room',)),'foreignOwnership':v(vals,('sohuu',),('ownership',))}
  else:z={'date':d,'propBuyValue':v(vals,('mua','giatri'),('mua','gt'),('buy','value')),'propSellValue':v(vals,('ban','giatri'),('ban','gt'),('sell','value')),'propNetValue':v(vals,('rong','giatri'),('rong','gt'),('net','value')),'propBuyVolume':v(vals,('mua','khoiluong'),('mua','kl'),('buy','volume')),'propSellVolume':v(vals,('ban','khoiluong'),('ban','kl'),('sell','volume')),'propNetVolume':v(vals,('rong','khoiluong'),('rong','kl'),('net','volume'))}
  out.append(z)
 fields=['foreignBuyValue','foreignSellValue','foreignNetValue'] if kind=='foreign' else ['propBuyValue','propSellValue','propNetValue']
 if out and not any(any(abs(float(x.get(k,0) or 0))>1e-12 for k in fields) for x in out):return []
 return sorted({x['date']:x for x in out}.values(),key=lambda x:x['date'])
def json_rows(sym,kind):
 ep='GDKhoiNgoai.ashx' if kind=='foreign' else 'GDTuDoanh.ashx';out=[]
 for pg in range(1,500):
  q=urlencode({'Symbol':sym,'StartDate':START,'EndDate':END,'PageIndex':pg,'PageSize':100})
  try:raw,_=get(BASE+ep+'?'+q,20);p=json.loads(raw.decode('utf-8','ignore'))
  except:break
  a=p.get('Data') or p.get('data') or p.get('DataRows') or []
  if isinstance(a,dict):a=a.get('Data') or a.get('Rows') or []
  if not isinstance(a,list) or not a:break
  for r in a:
   kk={nkey(k):v for k,v in r.items()};d=date_iso(next((v for k,v in kk.items() if 'ngay' in k or 'date' in k),None))
   if not d:continue
   def pick(*alts):
    for alt in alts:
     for k,vv in kk.items():
      if all(t in k for t in alt):return asnum(vv)
    return 0.
   if kind=='foreign':z={'date':d,'foreignBuyValue':pick(('mua','giatri'),('mua','gt'),('buy','value')),'foreignSellValue':pick(('ban','giatri'),('ban','gt'),('sell','value')),'foreignNetValue':pick(('rong','giatri'),('rong','gt'),('net','value')),'foreignBuyVolume':pick(('mua','khoiluong'),('mua','kl'),('buy','volume')),'foreignSellVolume':pick(('ban','khoiluong'),('ban','kl'),('sell','volume')),'foreignNetVolume':pick(('rong','khoiluong'),('rong','kl'),('net','volume')),'foreignRoom':pick(('room',)),'foreignOwnership':pick(('sohuu',),('ownership',))}
   else:z={'date':d,'propBuyValue':pick(('mua','giatri'),('mua','gt'),('buy','value')),'propSellValue':pick(('ban','giatri'),('ban','gt'),('sell','value')),'propNetValue':pick(('rong','giatri'),('rong','gt'),('net','value')),'propBuyVolume':pick(('mua','khoiluong'),('mua','kl'),('buy','volume')),'propSellVolume':pick(('ban','khoiluong'),('ban','kl'),('sell','volume')),'propNetVolume':pick(('rong','khoiluong'),('rong','kl'),('net','volume'))}
   out.append(z)
  total=int(p.get('TotalCount') or p.get('totalCount') or p.get('Total') or 0);pages=int(p.get('TotalPage') or p.get('totalPage') or p.get('TotalPages') or 0)
  if (pages and pg>=pages) or (total and len(out)>=total):break
  if len(a)<20 and not pages:break
 return sorted({x['date']:x for x in out}.values(),key=lambda x:x['date'])
def one(sym):
 audit={};parts={}
 for kind in ('foreign','prop'):
  try:r=export_rows(sym,kind);src='EXPORT_XLSX'
  except:r=[];src='EXPORT_ERROR'
  if len(r)<40:
   try:j=json_rows(sym,kind)
   except:j=[]
   if len(j)>len(r):r=j;src='JSON_FALLBACK'
  parts[kind]=r;audit[kind]=src;audit[kind+'Rows']=len(r)
 d={}
 for kind in ('foreign','prop'):
  for x in parts[kind]:d.setdefault(x['date'],{'date':x['date']}).update(x)
 return sym,[d[k] for k in sorted(d)],audit
def main():
 man=json.loads((ROOT/'data/hose-fallbacks/manifest.json').read_text(encoding='utf-8'));syms=[s for s,r in (man.get('routes') or {}).items() if int((r or {}).get('rows') or 0)>=520];allr={};aud={}
 with ThreadPoolExecutor(max_workers=10) as ex:
  fs={ex.submit(one,s):s for s in syms}
  for i,f in enumerate(as_completed(fs),1):
   s,r,a=f.result();allr[s]=r;aud[s]=a
   if i%50==0:print(json.dumps({'flowSymbolsDone':i,'total':len(syms)}))
 lens=[len(x) for x in allr.values()];fl=[a['foreignRows'] for a in aud.values()];pl=[a['propRows'] for a in aud.values()];cur={}
 for s,r in allr.items():
  if not r:continue
  z={}
  for typ in ('foreign','prop'):
   vv=np.asarray([float(x.get(typ+'NetValue',0) or 0) for x in r],float);gross=np.asarray([float(x.get(typ+'BuyValue',0) or 0)+float(x.get(typ+'SellValue',0) or 0) for x in r],float);nz=np.where(gross>0)[0]
   if len(nz)<20:continue
   j=int(nz[-1]);hist=vv[max(0,j-59):j+1];sd=float(hist.std(ddof=1)) if len(hist)>2 else 0.;z[typ+'Net1']=float(vv[j]);z[typ+'Net5']=float(vv[max(0,j-4):j+1].sum());z[typ+'Net20']=float(vv[max(0,j-19):j+1].sum());z[typ+'Z60']=float((vv[j]-hist.mean())/(sd or 1));gg=float(gross[max(0,j-19):j+1].sum());z[typ+'NetRatio20']=z[typ+'Net20']/gg if gg else 0.
  if z:
   z['date']=r[-1]['date'];last=r[-1]
   if 'foreignRoom' in last:z['foreignRoom']=last.get('foreignRoom');z['foreignOwnership']=last.get('foreignOwnership')
   cur[s]=z
 summary={'symbols':len(syms),'symbolsWithFlow':sum(x>0 for x in lens),'symbols100plus':sum(x>=100 for x in lens),'symbols500plus':sum(x>=500 for x in lens),'medianRows':int(np.median(lens)) if lens else 0,'p10Rows':int(np.quantile(lens,.1)) if lens else 0,'foreignMedianRows':int(np.median(fl)) if fl else 0,'propMedianRows':int(np.median(pl)) if pl else 0,'foreignExportSymbols':sum(a['foreign']=='EXPORT_XLSX' for a in aud.values()),'propExportSymbols':sum(a['prop']=='EXPORT_XLSX' for a in aud.values()),'currentForeign':sum('foreignNet1' in z for z in cur.values()),'currentProp':sum('propNet1' in z for z in cur.values())}
 out={'version':VERSION,'generatedAt':datetime.now(timezone.utc).isoformat(),'range':{'start':'2018-01-01','end':'2026-12-31'},'source':'CafeF historical foreign/proprietary trading export with deep JSON fallback','symbols':allr,'current':cur,'sourceAudit':aud,'summary':summary};(ROOT/'data/flow-v11.json').write_text(json.dumps(out,ensure_ascii=False,separators=(',',':')),encoding='utf-8');print(json.dumps(summary,ensure_ascii=False))
if __name__=='__main__':main()
